import streamlit as st
import pandas as pd
import numpy as np
import calendar
from datetime import date, timedelta, datetime
from io import BytesIO
import zipfile
import openpyxl
from copy import copy

# =============================
# Fonctions auxiliaires
# =============================

def get_all_days(mois, annee):
    jours = []
    current_date = date(annee, mois, 1)
    while current_date.month == mois:
        jours.append(current_date)
        current_date += timedelta(days=1)
    return jours

def get_jours_ouvres(mois, annee, jours_feries):
    all_days = get_all_days(mois, annee)
    jours_ouvres = [d for d in all_days if d.weekday() < 5 and d not in jours_feries]
    return jours_ouvres

def generer_excel(mois_selectionne, annee_selectionnee, contrats, heures_par_jour, jours_feries, donors=None):
    import openpyxl
    from copy import copy
    
    jours_mois = get_all_days(mois_selectionne, annee_selectionnee)
    jours_ouvres = get_jours_ouvres(mois_selectionne, annee_selectionnee, jours_feries)
    nb_jours_ouvres = len(jours_ouvres)
    HEURES_TOTALES = nb_jours_ouvres * heures_par_jour

    heures_cibles = {code: round(HEURES_TOTALES * pct / 100, 2) for code, pct in contrats.items()}
    contrats_list = list(contrats.keys())

    # Format dates as strings (YYYY-MM-DD)
    jours_mois_str = [d.strftime("%Y-%m-%d") for d in jours_mois]

    df_repartition = pd.DataFrame(index=contrats_list, columns=jours_mois_str, dtype=float)
    df_repartition[:] = 0.0
    heures_restantes = heures_cibles.copy()

    for jour, jour_str in zip(jours_mois, jours_mois_str):
        if jour.weekday() >= 5 or jour in jours_feries:
            df_repartition[jour_str] = np.nan
            continue
        max_alloc = [min(heures_restantes[code], heures_par_jour) for code in contrats_list]
        rng = np.random.default_rng()
        tries = 0
        while True:
            props = rng.dirichlet(np.ones(len(contrats_list)))
            alloc = np.minimum(np.round(props * heures_par_jour * 2) / 2, max_alloc)
            diff = heures_par_jour - alloc.sum()
            tries += 1
            if abs(diff) < 0.01:
                break
            idx = np.argmax(max_alloc)
            if alloc[idx] + diff <= max_alloc[idx] and alloc[idx] + diff >= 0:
                alloc[idx] += diff
                break
            # Add this line to see how many tries per day
            if tries > 1000:
                st.write(f"Warning: allocation for {jour_str} took {tries} tries")
                break
        for idx, code in enumerate(contrats_list):
            df_repartition.loc[code, jour_str] = alloc[idx]
            heures_restantes[code] -= alloc[idx]

    # Add Donor, Financing and Project columns to the left
    donor_values = [donors.get(code, "") if donors else "" for code in df_repartition.index]
    financing_values = list(df_repartition.index)
    project_values = ["" for code in df_repartition.index]  # Add your project data here
    df_repartition.insert(0, "", project_values)
    df_repartition.insert(0, "Financing Code", financing_values)
    df_repartition.insert(0, "Donor", donor_values)

    # Load the existing template and fill it with data
    try:
        # Load the template workbook
        template_wb = openpyxl.load_workbook("Trame timesheet.xlsx")
        
        # Create output in memory
        output = BytesIO()
        
        # Save the modified workbook to output
        template_wb.save(output)
        
        # Now reopen and fill with our data
        output.seek(0)
        wb = openpyxl.load_workbook(output)
        ws = wb.active
        
        # Fill the data starting from row 8 (as per your previous requirement)
        start_row = 8
        
        # Write headers
        for col_idx, col_name in enumerate(df_repartition.columns, start=1):
            ws.cell(row=start_row, column=col_idx, value=col_name)
        
        # Write data
        for row_idx, (index, row_data) in enumerate(df_repartition.iterrows(), start=start_row + 1):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
        
    except FileNotFoundError:
        st.warning("Template 'Trame timesheet.xlsx' not found. Creating new file...")
        # Fallback to original method if template not found
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_repartition.to_excel(writer, sheet_name="Planning", index=False, startrow=7)
        output.seek(0)
        return output

# =============================
# Language toggle (flags)
# =============================

from streamlit import markdown

LANGUAGES = {
    "Français": "🇫🇷",
    "English": "🇬🇧",
    "Español": "🇪🇸"
}

lang_labels = [f"{LANGUAGES[l]} {l}" for l in LANGUAGES]
lang_map = dict(zip(lang_labels, LANGUAGES.keys()))

st.markdown(
    """
    <style>
    .block-container {padding-top: 1rem;}
    </style>
    """,
    unsafe_allow_html=True
)

selected_lang_label = st.radio(
    label="Language Selection",
    options=lang_labels,
    horizontal=True,
    label_visibility="hidden"
)
lang = lang_map[selected_lang_label]
is_fr = lang == "Français"
is_en = lang == "English"
is_es = lang == "Español"

# =============================
# Interface Streamlit
# =============================

st.set_page_config(
    page_title=(
        "Générateur de Planning" if is_fr else
        "Timesheet Generator" if is_en else
        "Generador de Horarios"
    ),
    layout="centered"
)

st.title(
    "📅 Générateur de planning d'heures" if is_fr else
    "📅 Timesheet Generator" if is_en else
    "📅 Generador de horarios"
)

template = BytesIO()
df_template = pd.DataFrame({
    "Année" if is_fr else "Year" if is_en else "Año": [2025],
    "Mois" if is_fr else "Month" if is_en else "Mes": [10],
    "Heures par jour" if is_fr else "Hours per day" if is_en else "Horas por día": [8],
    "Jours fériés" if is_fr else "Holidays" if is_en else "Días festivos": ["2025-10-01,2025-10-15"],
    "Contrats" if is_fr else "Contracts" if is_en else "Contratos": ["Contract1:50,Contract2:50"],
    "Bailleurs" if is_fr else "Donors" if is_en else "Donarios": ["Donor1,Donor2"]
})
with pd.ExcelWriter(template, engine="openpyxl") as writer:
    df_template.to_excel(writer, index=False)
template.seek(0)

st.markdown(
    (
        "ℹ️ [Comment utiliser le modèle Excel ?](#)<br>"
        "Cliquez sur le bouton ci-dessous pour télécharger un modèle Excel.<br>"
        "Remplissez chaque ligne avec vos paramètres (année, mois, heures par jour, jours fériés, contrats, donneurs).<br>"
        "Ensuite, importez ce fichier pour générer automatiquement tous vos plannings."
        if is_fr else
        "ℹ️ [How to use the Excel template?](#)<br>"
        "Click the button below to download an Excel template.<br>"
        "Fill each row with your parameters (year, month, hours per day, holidays, contracts, donors).<br>"
        "Then, upload this file to automatically generate all your timesheets."
        if is_en else
        "ℹ️ [¿Cómo usar la plantilla de Excel?](#)<br>"
        "Haga clic en el botón de atras para descargar una plantilla de Excel.<br>"
        "Complete cada fila con sus parámetros (año, mes, horas por día, días festivos, contratos, donantes).<br>"
        "Luego, suba este archivo para generar automáticamente todos sus horarios."
    ),
    unsafe_allow_html=True
)

st.download_button(
    label=(
        "📥 Télécharger le modèle Excel" if is_fr else
        "📥 Download Excel template" if is_en else
        "📥 Descargar plantilla Excel"
    ),
    data=template,
    file_name=(
        "modele_plannings.xlsx" if is_fr else
        "timesheet_template.xlsx" if is_en else
        "plantilla_horarios.xlsx"
    ),
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


# Upload multiple plannings
st.subheader(
    "Importer un fichier Excel pour générer les plannings annuels" if is_fr else
    "Upload an Excel file to generate annual timesheets" if is_en else
    "Subir un archivo Excel para generar los horarios anuales"
)
uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"], label_visibility="hidden")

if uploaded_file:
    df_upload = pd.read_excel(uploaded_file)
    st.write(
        "Aperçu du fichier importé :" if is_fr else
        "Preview of uploaded file:" if is_en else
        "Vista previa del archivo subido:"
    )
    st.dataframe(df_upload)

    # Initialize session state for download files
    if 'zip_data' not in st.session_state:
        st.session_state.zip_data = None
    
    if st.button(
        "✅ Générer tous les plannings du fichier" if is_fr else
        "✅ Generate all timesheets from file" if is_en else
        "✅ Generar todos los horarios del archivo"
    ):
        year_col = "Année" if is_fr else "Year" if is_en else "Año"
        grouped = df_upload.groupby(year_col)
        download_files = []

        progress_bar = st.progress(0)
        total_rows = len(df_upload)
        processed_rows = 0

        for year, group in grouped:
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                sheet_written = False
                for idx, row in group.iterrows():
                    try:
                        mois = int(row["Mois"] if is_fr else row["Month"] if is_en else row["Mes"])
                        heures_par_jour = int(row["Heures par jour"] if is_fr else row["Hours per day"] if is_en else row["Horas por día"])
                        jours_feries = []
                        jours_feries_col = "Jours fériés" if is_fr else "Holidays" if is_en else "Días festivos"
                        if pd.notna(row.get(jours_feries_col, None)):
                            for d in str(row[jours_feries_col]).split(","):
                                d = d.strip()
                                if d:
                                    jours_feries.append(datetime.strptime(d, "%Y-%m-%d").date())
                        contrats = {}
                        donors = {}
                        contrats_col = "Contrats" if is_fr else "Contracts" if is_en else "Contratos"
                        donor_col = "Bailleurs" if is_fr else "Donors" if is_en else "Donarios"
                        contrats_items = str(row[contrats_col]).split(",")
                        donor_items = str(row.get(donor_col, "")).split(",")
                        if len(donor_items) != len(contrats_items):
                            st.warning(
                                f"Ligne {idx+1} ignorée : nombre de donors ({len(donor_items)}) différent du nombre de contrats ({len(contrats_items)})" if is_fr else
                                f"Row {idx+1} skipped: number of donors ({len(donor_items)}) does not match number of contracts ({len(contrats_items)})" if is_en else
                                f"Fila {idx+1} omitida: número de donantes ({len(donor_items)}) diferente al número de contratos ({len(contrats_items)})"
                            )
                            continue
                        for i, item in enumerate(contrats_items):
                            code, pct = item.split(":")
                            contrats[code.strip()] = float(pct.strip())
                            donors[code.strip()] = donor_items[i].strip()
                        st.write(f"Contrats parsed: {contrats}, sum: {sum(contrats.values())}")
                        if sum(contrats.values()) != 100:
                            st.warning(
                                f"Ligne {idx+1} ignorée : la somme des pourcentages de contrats n'est pas 100 (somme: {sum(contrats.values())})" if is_fr else
                                f"Row {idx+1} skipped: contract percentages do not sum to 100 (sum: {sum(contrats.values())})" if is_en else
                                f"Fila {idx+1} omitida: los porcentajes de contratos no suman 100 (suma: {sum(contrats.values())})"
                            )
                            continue
                        excel_file = generer_excel(mois, year, contrats, heures_par_jour, jours_feries, donors)
                        # Read the generated workbook and copy it to our output
                        temp_wb = openpyxl.load_workbook(excel_file)
                        temp_ws = temp_wb.active
                        
                        # Create new worksheet in our output with month name
                        sheet_name = f"{calendar.month_name[mois]}"
                        new_ws = writer.book.create_sheet(title=sheet_name)
                        
                        # Copy all data and formatting from template to new sheet
                        for row in temp_ws.iter_rows():
                            for cell in row:
                                new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                                # Copy formatting if needed
                                if cell.has_style:
                                    new_cell.font = copy(cell.font)
                                    new_cell.border = copy(cell.border)
                                    new_cell.fill = copy(cell.fill)
                                    new_cell.number_format = copy(cell.number_format)
                                    new_cell.protection = copy(cell.protection)
                                    new_cell.alignment = copy(cell.alignment)
                        
                        sheet_written = True
                    except Exception as e:
                        st.warning(
                            f"Ligne {idx+1} ignorée : {e}" if is_fr else
                            f"Row {idx+1} skipped: {e}" if is_en else
                            f"Fila {idx+1} omitida: {e}"
                        )
                    processed_rows += 1
                    progress_bar.progress(processed_rows / total_rows)
                if not sheet_written:
                    pd.DataFrame({"Info": ["No valid rows"]}).to_excel(writer, sheet_name="Info", index=False)
            output.seek(0)
            download_files.append((year, output))

        st.success(
            "Tous les plannings ont été générés !" if is_fr else
            "All timesheets have been generated!" if is_en else
            "¡Todos los horarios han sido generados!"
        )

        # Create ZIP file and store in session state
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
            for year, fileobj in download_files:
                filename = (
                    f"plannings_{year}.xlsx" if is_fr else
                    f"timesheets_{year}.xlsx" if is_en else
                    f"horarios_{year}.xlsx"
                )
                zipf.writestr(filename, fileobj.getvalue())
        zip_buffer.seek(0)
        
        # Store zip data in session state
        st.session_state.zip_data = zip_buffer.getvalue()

    # Show download button only if zip data is available
    if st.session_state.zip_data is not None:
        st.download_button(
            label=(
                "📥 Télécharger tous les plannings (ZIP)" if is_fr else
                "📥 Download all timesheets (ZIP)" if is_en else
                "📥 Descargar todos los horarios (ZIP)"
            ),
            data=st.session_state.zip_data,
            file_name=(
                "plannings_annuels.zip" if is_fr else
                "yearly_timesheets.zip" if is_en else
                "horarios_anuales.zip"
            ),
            mime="application/zip"
        )


