import pandas as pd
import openpyxl

try:
    # Load the Excel template using openpyxl to preserve formatting
    wb = openpyxl.load_workbook('Trame timesheet.xlsx')
    
    print("Worksheets in the template:")
    for sheet_name in wb.sheetnames:
        print(f"- {sheet_name}")
        
    # Get the first worksheet
    ws = wb.active
    print(f"\nActive sheet: {ws.title}")
    
    # Read a reasonable range to see the structure
    print(f"\nSheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
    
    print("\nFirst 15 rows and columns:")
    for row in range(1, min(16, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(16, ws.max_column + 1)):
            cell_value = ws.cell(row=row, column=col).value
            row_data.append(str(cell_value) if cell_value is not None else "")
        print(f"Row {row}: {row_data}")
        
    # Also try reading with pandas to see data structure
    print("\n" + "="*50)
    print("Reading with pandas:")
    df = pd.read_excel('Trame timesheet.xlsx')
    print("Columns:", df.columns.tolist())
    print("Shape:", df.shape)
    print("\nFirst few rows:")
    print(df.head())
    
except Exception as e:
    print(f"Error: {e}")